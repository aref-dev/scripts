from openpyxl import Workbook, load_workbook
import os

# File path to the Excel file
file_path = "practice.xlsx"

# The text to append
text = """
Lorem ipsum dolor sit amet, consectetuer adipiscing elit.
Aliquam tincidunt mauris eu risus.
Vestibulum auctor dapibus neque.
Nunc dignissim risus id metus.
Cras ornare tristique elit.
Vivamus vestibulum ntulla nec ante.
"""

# Check if the file exists
if not os.path.exists(file_path):
    # Create a new workbook and save it
    wb = Workbook()
    ws = wb.active
    wb.save(file_path)

# Load the workbook and select the active sheet
wb = load_workbook(file_path)
ws = wb.active

# Scenario 1: Append to the next available row
next_row = ws.max_row + 1  # Automatically determines the next available row
start_column = 1  # Starting column (e.g., C corresponds to column 3)
# Insert the text into the next available row and specified column
ws.cell(row=next_row, column=start_column, value=text)

# OR Scenario 2: Append to a specified row
# specified_row = 10  # Change this to your desired row
# ws.cell(row=specified_row, column=start_column, value=text)

# Save the workbook
wb.save(file_path)
